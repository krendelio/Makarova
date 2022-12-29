import csv
import datetime
import sys
from datetime import datetime
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import numpy as np
import matplotlib.pyplot as plt
from jinja2 import Environment, FileSystemLoader
import pdfkit

currency_to_rub = {'AZN': 35.68, 'BYR': 23.91, 'EUR': 59.90, 'GEL': 21.74, 'KGS': 0.76, 'KZT': 0.13, 'RUR': 1,
                   'UAH': 1.64, 'USD': 60.66, 'UZS': 0.0055}

class DataSet:
    def __init__(self, file_name):
        self.reader = [row for row in csv.reader(open(file_name, encoding='utf_8_sig'))]
        self.reader_conditions()

    def reader_conditions(self):
        if len(self.reader) == 0:
            print('Пустой файл')
            sys.exit()
        self.columns_names = self.reader[0]
        self.vacancies_data = [row for row in self.reader[1:] if
                               len(row) == len(self.columns_names) and row.count('') == 0]
        if len(self.vacancies_data) == 0:
            print('Нет данных')
            sys.exit()


class UsersInput:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.profession_name = input('Введите название профессии: ')
        self.file_name = self.check_file(self.file_name)
        self.profession_name = self.check_profession(self.profession_name)

    @staticmethod
    def check_profession(profession_name):
        if profession_name == '':
            print('Некорректное название профессии')
            sys.exit()
        return profession_name

    @staticmethod
    def check_file(file_name):
        if file_name == '' or '.' not in file_name:
            print('Некорректное название файла')
            sys.exit()
        return file_name


class Vacancy:
    name: str
    salary_from: int or float
    salary_to: int or float
    salary_currency: str
    area_name: str
    published_at: str
    salary: str

    def __init__(self, vacancy):
        for key, value in vacancy.items():
            self.__setattr__(key, formatter(key, value))


def formatter(key, value):
    if key in ['salary_from', 'salary_to']:
        return float(value)
    if key == 'published_at':
        return int(datetime.strptime(value, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))
    return value


class SalaryDict:
    def __init__(self):
        self.salary_dict = {}
        self.__aver_salary_dict = {}

    def add_salary(self, key, salary):
        if self.salary_dict.get(key) is None:
            self.salary_dict[key] = []
        return self.salary_dict[key].append(salary)

    def get_aver_salary(self):
        for key, value in self.salary_dict.items():
            self.__aver_salary_dict[key] = int(mean(value))
        return self.__aver_salary_dict


class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency


class DictCount:
    def __init__(self):
        self.length = 0
        self.count_dict = {}
        self.big_cities = []
        self.top_proportion_dict = {}

    def dictAdd(self, key):
        if self.count_dict.get(key) is None:
            self.count_dict[key] = 0
        self.count_dict[key] += 1
        self.length += 1
        return

    def proportion(self):
        proportion_dict = {}
        for key, value in self.count_dict.items():
            proportion = value / self.length
            self.proportion_condition(key, proportion, proportion_dict)
        sorted_dict = dict(sorted(proportion_dict.items(), key=lambda row: row[1], reverse=True))
        self.top_proportion_dict = {x: sorted_dict[x] for x in list(sorted_dict)[:10]}
        return

    def proportion_condition(self, key, proportion, proportion_dict):
        if proportion >= 0.1:
            self.big_cities.append(key)
            proportion_dict[key] = round(proportion, 4)


class ParseData:
    def __init__(self):
        self.salary_lvl_by_year = SalaryDict()
        self.count_vac_by_year = DictCount()
        self.salary_lvl_by_year_for_prof = SalaryDict()
        self.count_vac_by_year_for_prof = DictCount()
        self.salary_lvl_by_city = SalaryDict()
        self.vacancy_rate_by_city = DictCount()

    def get_data(self, vacancies, prof):
        for vacancy in vacancies:
            vacancy_salary = (vacancy.salary_from + vacancy.salary_to) / 2 * currency_to_rub[vacancy.salary_currency]
            self.salary_lvl_by_year.add_salary(vacancy.published_at, vacancy_salary)
            self.count_vac_by_year.dictAdd(vacancy.published_at)
            self.salary_lvl_by_city.add_salary(vacancy.area_name, vacancy_salary)
            self.vacancy_rate_by_city.dictAdd(vacancy.area_name)
            if prof in vacancy.name:
                self.salary_lvl_by_year_for_prof.add_salary(vacancy.published_at, vacancy_salary)
                self.count_vac_by_year_for_prof.dictAdd(vacancy.published_at)
        self.condition()
        self.salary_lvl_by_city, list_del_town = self.get_top_salary(self.salary_lvl_by_city)
        self.vacancy_rate_by_city.proportion()
        self.vacancy_rate_by_city = self.get_rate_by_city(self.vacancy_rate_by_city)
        self.vacancy_rate_by_city = dict((x, y) for x, y in self.vacancy_rate_by_city)
        return self.salary_lvl_by_year.get_aver_salary(), self.count_vac_by_year.count_dict, \
               self.salary_lvl_by_year_for_prof.get_aver_salary(), self.count_vac_by_year_for_prof.count_dict, \
               self.salary_lvl_by_city, self.vacancy_rate_by_city

    def condition(self):
        if self.salary_lvl_by_year_for_prof.salary_dict == {}:
            self.salary_lvl_by_year_for_prof.salary_dict = {x: [0] for x in self.salary_lvl_by_year.salary_dict.keys()}
        elif self.salary_lvl_by_year_for_prof.salary_dict != {} and len(
                list(self.salary_lvl_by_year.get_aver_salary().keys())) != len(
            list(self.salary_lvl_by_year_for_prof.get_aver_salary().keys())):
            for key in list(self.salary_lvl_by_year.get_aver_salary().keys()):
                if key not in list(self.salary_lvl_by_year_for_prof.get_aver_salary().keys()):
                    self.salary_lvl_by_year_for_prof.get_aver_salary()[key] = 0
        if self.count_vac_by_year_for_prof.count_dict == {}:
            self.count_vac_by_year_for_prof.count_dict = {x: 0 for x in self.count_vac_by_year.count_dict.keys()}
        elif self.count_vac_by_year_for_prof.count_dict != {} and len(
                list(self.count_vac_by_year.count_dict.keys())) != len(
            list(self.count_vac_by_year_for_prof.count_dict.keys())):
            for key in list(self.count_vac_by_year.count_dict.keys()):
                if key not in list(self.count_vac_by_year_for_prof.count_dict.keys()):
                    self.count_vac_by_year_for_prof.count_dict[key] = 0

    @staticmethod
    def get_top_salary(list_all_salary):
        average = []
        town_count = {}
        for i in range(len(list_all_salary.salary_dict)):
            town = list(list_all_salary.salary_dict)[i]
            s = list(list_all_salary.salary_dict.values())[i]
            town_count[town] = len(s)
            aver = int(sum(s) / len(s))
            average.append((town, aver))

        s = sum(town_count.values())
        list_del_town = []
        list_del_town_index = []
        ParseData.get_salary(average, list_del_town, list_del_town_index, s, town_count)

        top_salary = dict(sorted(average, key=lambda row: row[1], reverse=True))
        salary_dict = {}
        for key, value in top_salary.items():
            salary_dict[key] = value
        return {x: salary_dict[x] for x in list(salary_dict)[:10]}, list_del_town

    @staticmethod
    def get_salary(average, list_del_town, list_del_town_index, s, town_count):
        for i in range(len(town_count.items())):
            n = list(town_count.values())[i]
            town = list(town_count)[i]
            percent = round(100 * int(n) / s, 1)
            if percent < 1 or town == 'Россия':
                list_del_town.append((town, n))
                list_del_town_index.append(i)
        for i in reversed(range(len(list_del_town))):
            del town_count[list_del_town[i][0]]
            del average[list_del_town_index[i]]

    @staticmethod
    def get_rate_by_city(vacancy_rate_by_city):
        s = vacancy_rate_by_city.length
        del_town = []
        for i in reversed(range(len(del_town))):
            del vacancy_rate_by_city.count_dict[del_town[i][0]]

        proportion_dict = {}
        for key, value in vacancy_rate_by_city.count_dict.items():
            proportion = value / s
            if proportion >= 0.01:
                proportion_dict[key] = round(proportion, 4)

        sorted_dict = sorted(proportion_dict.items(), key=lambda row: row[1], reverse=True)
        return sorted_dict[:10]


class Report:
    def __init__(self, data, prof):
        self.salary_lvl_by_year = data[0]
        self.count_vac_by_year = data[1]
        self.salary_lvl_by_year_for_prof = data[2]
        self.count_vac_by_year_for_prof = data[3]
        self.salary_lvl_by_city = data[4]
        self.vacancy_rate_by_city = data[5]
        self.prof = prof

        self.wb = Workbook()
        self.sheet1 = self.wb.active
        self.sheet1.title = 'Статистика по годам'
        self.sheet2 = self.wb.create_sheet('Статистика по городам')

        self.fig = plt.figure()
        self.ax1 = self.fig.add_subplot(221)
        self.ax1.set_title('Уровень зарплат по годам')
        self.ax2 = self.fig.add_subplot(222)
        self.ax2.set_title('Количество вакансий по годам')
        self.ax3 = self.fig.add_subplot(223)
        self.ax3.set_title('Уровень зарплат по городам')
        self.ax4 = self.fig.add_subplot(224)
        self.ax4.set_title('Доля вакансий по городам')

    def generate_excel(self):
        names_sheet1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.prof}',
                        'Количество вакансий', f'Количество вакансий - {self.prof}']
        names_sheet2 = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']

        self.excel_conditions(names_sheet1, names_sheet2)

        side = Side(border_style='thin', color='000000')
        self.set_border(self.sheet1, side)
        self.set_border(self.sheet2, side)
        self.sheet2.insert_cols(3)
        self.sheet2.column_dimensions['C'].width = 2

        self.column_width(self.sheet1)
        self.column_width(self.sheet2)

        for i in range(2, len(self.sheet2['E']) + 1):
            self.sheet2[f'E{i}'].number_format = FORMAT_PERCENTAGE_00

        self.wb.save('report.xlsx')

    def excel_conditions(self, names_sheet1, names_sheet2):
        for i, name in enumerate(names_sheet1):
            self.sheet1.cell(row=1, column=(i + 1), value=name).font = Font(bold=True)
        for year, value in self.salary_lvl_by_year.items():
            self.sheet1.append([year, value, self.salary_lvl_by_year_for_prof[year], self.count_vac_by_year[year],
                                self.count_vac_by_year_for_prof[year]])
        for i, name in enumerate(names_sheet2):
            self.sheet2.cell(row=1, column=(i + 1), value=name).font = Font(bold=True)
        for i in range(len(list(self.salary_lvl_by_city.keys()))):
            self.sheet2.append([list(self.salary_lvl_by_city.keys())[i], list(self.salary_lvl_by_city.values())[i],
                                list(self.vacancy_rate_by_city.keys())[i], list(self.vacancy_rate_by_city.values())[i]])

    @staticmethod
    def set_border(ws, side):
        for cell in ws._cells.values():
            cell.border = Border(top=side, bottom=side, left=side, right=side)

    @staticmethod
    def column_width(ws):
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value + 2

    def image_generate(self):
        width_12 = 0.4
        x_nums_1 = np.arange(len(self.salary_lvl_by_year.keys()))
        x_list1_1 = x_nums_1 - width_12 / 2
        x_list1_2 = x_nums_1 + width_12 / 2
        self.image(width_12, x_list1_1, x_list1_2, x_nums_1)

        other = 1
        data = [1]
        labels = ['Другие']
        for key, value in self.vacancy_rate_by_city.items():
            data.append(value * 100)
            labels.append(key)
            other -= value
        data[0] = round(other, 4) * 100
        textprops = {"fontsize": 6}

        self.ax4.pie(data, labels=labels, textprops=textprops, radius=1.1)
        plt.tight_layout()
        plt.savefig('graph.png')

    def image(self, width_12, x_list1_1, x_list1_2, x_nums_1):
        self.ax1.bar(x_list1_1, self.salary_lvl_by_year.values(), width_12, label='средняя з/п')
        self.ax1.bar(x_list1_2, self.salary_lvl_by_year_for_prof.values(), width_12, label=f'з/п {self.prof}')
        self.ax1.set_xticks(x_nums_1, self.salary_lvl_by_year.keys(), rotation='vertical')
        self.ax1.tick_params(axis='both', labelsize=8)
        self.ax1.legend(fontsize=8)
        self.ax1.grid(True, axis='y')
        x_nums_2 = np.arange(len(self.count_vac_by_year.keys()))
        x_list2_1 = x_nums_2 - width_12 / 2
        x_list2_2 = x_nums_2 + width_12 / 2
        self.ax2.bar(x_list2_1, self.count_vac_by_year.values(), width_12, label='Количество вакансий')
        self.ax2.bar(x_list2_2, self.count_vac_by_year_for_prof.values(), width_12,
                     label=f'Количество вакансий\n{self.prof}')
        self.ax2.set_xticks(x_nums_2, self.count_vac_by_year.keys(), rotation='vertical')
        self.ax2.tick_params(axis='both', labelsize=8)
        self.ax2.legend(fontsize=8)
        self.ax2.grid(True, axis='y')
        list_names = self.image_conditions()
        width_3 = 0.7
        y_nums = np.arange(len(list(list_names.keys())))
        self.ax3.barh(y_nums, list_names.values(), width_3, align='center')
        self.ax3.set_yticks(y_nums, list_names.keys())
        self.ax3.tick_params(axis='y', labelsize=6)
        self.ax3.tick_params(axis='x', labelsize=8)
        self.ax3.invert_yaxis()
        self.ax3.grid(True, axis='x')

    def image_conditions(self):
        list_names = {}
        for key, value in self.salary_lvl_by_city.items():
            if ' ' in key:
                key = str(key).replace(' ', '\n')
            elif '-' in key and key.count('-') == 1:
                key = str(key).replace('-', '-\n')
            elif '-' in key and key.count('-') != 1:
                key = str(key).replace('-', '-\n', 1)
            list_names[key] = value
        return list_names

    def generate_pdf(self):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template('pdf_template.html')
        names_sheet1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.prof}',
                        'Количество вакансий', f'Количество вакансий - {self.prof}']
        names_sheet2 = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']
        statistics_by_city_dic = []
        pdf_template = self.pdf_conditions(names_sheet1, names_sheet2, statistics_by_city_dic, template)
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

    def pdf_conditions(self, names_sheet1, names_sheet2, statistics_by_city_dic, template):
        for i in range(len(list(self.salary_lvl_by_city.keys()))):
            statistics_by_city_dic.append(
                [list(self.salary_lvl_by_city.keys())[i], list(self.vacancy_rate_by_city.keys())[i]])
        for key, value in self.vacancy_rate_by_city.items():
            self.vacancy_rate_by_city[key] = str(round(value * 100, 2)) + '%'
        pdf_template = template.render({'name': self.prof,
                                        'salary_lvl_by_year': self.salary_lvl_by_year,
                                        'count_vac_by_year': self.count_vac_by_year,
                                        'salary_lvl_by_year_for_prof': self.salary_lvl_by_year_for_prof,
                                        'count_vac_by_year_for_prof': self.count_vac_by_year_for_prof,
                                        'statistics_by_city_dic': statistics_by_city_dic,
                                        'salary_lvl_by_city': self.salary_lvl_by_city,
                                        'vacancy_rate_by_city': self.vacancy_rate_by_city,
                                        'names_sheet1': names_sheet1,
                                        'names_sheet2': names_sheet2})
        return pdf_template


def output(data_vacancies, profession_name):
    all_vacancies = []
    for data_vacancy in data_vacancies:
        data_vacancy = Vacancy(dict(zip(column_names, data_vacancy)))
        all_vacancies.append(data_vacancy)
    data = ParseData()
    data = data.get_data(all_vacancies, profession_name)
    print(f'Динамика уровня зарплат по годам: {data[0]}')
    print(f'Динамика количества вакансий по годам: {data[1]}')
    print(f'Динамика уровня зарплат по годам для выбранной профессии: {data[2]}')
    print(f'Динамика количества вакансий по годам для выбранной профессии: {data[3]}')
    print(f'Уровень зарплат по городам (в порядке убывания): {data[4]}')
    print(f'Доля вакансий по городам (в порядке убывания): {data[5]}')
    return data


input = UsersInput()
dataset = DataSet(input.file_name)
(column_names, vacancies_data) = dataset.columns_names, dataset.vacancies_data
data_output = output(vacancies_data, input.profession_name)
final = Report(data_output, input.profession_name)
final.generate_excel()
final.image_generate()
final.generate_pdf()


